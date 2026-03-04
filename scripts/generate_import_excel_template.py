#!/usr/bin/env python
"""
Genera plantilla Excel para importación de transacciones.

Ejecutar: poetry run python -m scripts.generate_import_excel_template

El archivo se guarda en: templates/plantilla_importacion_transacciones.xlsx

Columnas alineadas al JSON de POST /transactions/import/ para que el frontend
pueda parsear y enviar correctamente.
"""
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Importación"

    # Encabezados (una fila por sección para claridad)
    headers = [
        # Usuario origen
        "user_origin_names",
        "user_origin_lastnames",
        "user_origin_email",
        "user_origin_password",
        # Cuenta bancaria origen
        "bank_account_origin_bank_id",
        "bank_account_origin_account_flow",
        "bank_account_origin_account_holder_type",
        "bank_account_origin_bank_country",
        "bank_account_origin_holder_names",
        "bank_account_origin_holder_surnames",
        "bank_account_origin_document_number",
        "bank_account_origin_business_name",
        "bank_account_origin_ruc_number",
        "bank_account_origin_account_number",
        "bank_account_origin_cci_number",
        "bank_account_origin_pix_key",
        "bank_account_origin_pix_key_type",
        "bank_account_origin_cpf",
        # Usuario destino
        "user_destination_names",
        "user_destination_lastnames",
        "user_destination_email",
        "user_destination_password",
        # Cuenta bancaria destino
        "bank_account_destination_bank_id",
        "bank_account_destination_account_flow",
        "bank_account_destination_account_holder_type",
        "bank_account_destination_bank_country",
        "bank_account_destination_holder_names",
        "bank_account_destination_holder_surnames",
        "bank_account_destination_document_number",
        "bank_account_destination_business_name",
        "bank_account_destination_ruc_number",
        "bank_account_destination_account_number",
        "bank_account_destination_cci_number",
        "bank_account_destination_pix_key",
        "bank_account_destination_pix_key_type",
        "bank_account_destination_cpf",
        # Transacción
        "tax_rate_id",
        "commission_id",
        "status",
        "origin_amount",
        "destination_amount",
        "commission_result",
        "total_to_send",
        "coupon_id",
        "send_date",
        "payment_date",
    ]

    # Fila 1: encabezados
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Fila 2: descripciones/ayuda para el usuario
    descriptions = [
        "Nombres emisor",
        "Apellidos emisor",
        "Email emisor",
        "Contraseña emisor (min 8 chars, mayúsc, minúsc, número, especial)",
        "UUID bank origen",
        "origin",
        "naturalPerson | legalEntity | generalAspect",
        "pe | br",
        "Titular nombres",
        "Titular apellidos",
        "DNI/CE",
        "Razón social (si persona jurídica)",
        "RUC (si aplica)",
        "Número cuenta",
        "CCI",
        "Clave PIX",
        "cpf | email | phone | random",
        "CPF (Brasil)",
        "Nombres receptor",
        "Apellidos receptor",
        "Email receptor",
        "Contraseña receptor",
        "UUID bank destino",
        "destination",
        "naturalPerson | legalEntity | generalAspect",
        "pe | br",
        "Titular nombres",
        "Titular apellidos",
        "DNI/CE",
        "Razón social",
        "RUC",
        "Número cuenta",
        "CCI",
        "Clave PIX",
        "cpf | email | phone | random",
        "CPF",
        "UUID tax_rate",
        "UUID commission",
        "pending | completed | failed",
        "Monto origen",
        "Monto destino",
        "Comisión",
        "Total a enviar",
        "UUID cupón (opcional)",
        "YYYY-MM-DD",
        "YYYY-MM-DD",
    ]
    for col, d in enumerate(descriptions, 1):
        ws.cell(row=2, column=col, value=d)
        ws.cell(row=2, column=col).font = Font(italic=True, color="666666")

    # Fila 3: ejemplo de datos
    example_row = [
        "Juan",
        "Pérez",
        "juan@example.com",
        "MiClave123!",
        "",  # bank_id - rellenar UUID
        "origin",
        "naturalPerson",
        "pe",
        "Juan",
        "Pérez",
        "12345678",
        "",
        "",
        "00123456789",
        "",
        "",
        "",
        "",
        "María",
        "García",
        "maria@example.com",
        "OtraClave456!",
        "",  # bank_id - rellenar UUID
        "destination",
        "naturalPerson",
        "br",
        "María",
        "García",
        "98765432",
        "",
        "",
        "00123456789",
        "",
        "",
        "",
        "",
        "",  # tax_rate_id
        "",  # commission_id
        "pending",
        "1000",
        "950",
        "50",
        "1000",
        "",
        "2025-02-24",
        "2025-02-24",
    ]
    for col, v in enumerate(example_row, 1):
        ws.cell(row=3, column=col, value=v)

    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Congelar fila de encabezados
    ws.freeze_panes = "A2"

    # Guardar
    out_dir = project_root / "templates"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "plantilla_importacion_transacciones.xlsx"
    wb.save(out_path)
    print(f"Plantilla guardada en: {out_path}")


if __name__ == "__main__":
    main()
